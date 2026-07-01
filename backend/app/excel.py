from openpyxl import Workbook, load_workbook

from . import config, models

HEADERS = ["id", "거래처", "납품일", "차량번호", "송장번호", "품명", "규격", "단위", "수량", "중량", "비고"]


def append_invoice(invoice: models.Invoice) -> None:
    if config.EXCEL_PATH.exists():
        workbook = load_workbook(config.EXCEL_PATH)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    sheet_name = (invoice.material_type or "미분류")[:31]
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(HEADERS)
    else:
        sheet = workbook[sheet_name]

    sheet.append([
        invoice.id,
        invoice.vendor,
        invoice.delivery_date.isoformat() if invoice.delivery_date else "",
        invoice.vehicle_no,
        invoice.invoice_no,
        invoice.item_name,
        invoice.spec,
        invoice.unit,
        invoice.quantity,
        invoice.weight,
        invoice.note,
    ])
    workbook.save(config.EXCEL_PATH)
