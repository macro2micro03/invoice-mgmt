from copy import copy as copy_style
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from . import report_photos

CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "material_inspection_form.xlsx"

MATERIAL_ROW_START = 9
MATERIAL_ROW_END = 24
MATERIAL_ROW_CAPACITY = MATERIAL_ROW_END - MATERIAL_ROW_START + 1

WORK_TYPE_CELL = "B3"
CHECKLIST_RESULT_ROWS = range(63, 80)

PHOTO_SET_ROW_START = 81
PHOTO_SET_BLOCK_ROWS = 6
MAX_PHOTO_SETS = 5


def _mark_work_type_checkbox(text: str, work_type: str) -> str:
    return text.replace(f"{work_type} □", f"{work_type} ■", 1)


def _format_ton(value: float) -> str:
    return f"{value:.3f}".rstrip("0").rstrip(".")


def _format_korean_date(value: date) -> str:
    return f"{value:%Y}년 {value:%m}월 {value:%d}일"


def _build_material_spec_summary(material_type: str, specs: list[dict]) -> str:
    if not specs:
        return material_type
    first_spec = specs[0]["spec"]
    remaining = len(specs) - 1
    if remaining <= 0:
        return f"{material_type} {first_spec}"
    return f"{material_type} {first_spec} 외 {remaining}"


def _copy_photo_set_block(sheet, source_start: int, target_start: int) -> None:
    row_offset = target_start - source_start
    for offset in range(PHOTO_SET_BLOCK_ROWS):
        src_row = source_start + offset
        dst_row = target_start + offset
        sheet.row_dimensions[dst_row].height = sheet.row_dimensions[src_row].height
        for col in range(1, 11):
            src_cell = sheet.cell(row=src_row, column=col)
            dst_cell = sheet.cell(row=dst_row, column=col)
            dst_cell.value = src_cell.value
            dst_cell.font = copy_style(src_cell.font)
            dst_cell.border = copy_style(src_cell.border)
            dst_cell.fill = copy_style(src_cell.fill)
            dst_cell.alignment = copy_style(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= source_start and merged_range.max_row < source_start + PHOTO_SET_BLOCK_ROWS:
            min_col_letter = get_column_letter(merged_range.min_col)
            max_col_letter = get_column_letter(merged_range.max_col)
            new_min_row = merged_range.min_row + row_offset
            new_max_row = merged_range.max_row + row_offset
            sheet.merge_cells(f"{min_col_letter}{new_min_row}:{max_col_letter}{new_max_row}")


def fill_material_inspection_form(
    template_path,
    *,
    project_name: str,
    work_type: str,
    material_type: str,
    document_number: str,
    sender: str,
    receiver: str,
    checklist_sender: str,
    checklist_supervisor: str,
    specs: list[dict],
    vendor: str,
    delivery_date: str,
    photo_sets: list[dict] | None = None,
) -> tuple[bytes, list[dict]]:
    workbook = load_workbook(template_path)
    sheet = workbook.active

    today_date = date.today()
    today = today_date.strftime("%Y-%m-%d")
    today_korean = _format_korean_date(today_date)

    sheet["B2"] = project_name
    sheet[WORK_TYPE_CELL] = _mark_work_type_checkbox(str(sheet[WORK_TYPE_CELL].value or ""), work_type)
    sheet["B4"] = document_number
    sheet["G4"] = today_korean
    sheet["G5"] = today_korean

    fillable_specs = specs[:MATERIAL_ROW_CAPACITY]
    skipped_specs = specs[MATERIAL_ROW_CAPACITY:]
    for offset, spec_row in enumerate(fillable_specs):
        row = MATERIAL_ROW_START + offset
        sheet[f"A{row}"] = material_type
        sheet[f"A{row}"].alignment = CENTER_ALIGNMENT
        sheet[f"B{row}"] = spec_row["spec"]
        sheet[f"B{row}"].alignment = CENTER_ALIGNMENT
        sheet[f"D{row}"] = "Ton"
        sheet[f"D{row}"].alignment = CENTER_ALIGNMENT
        sheet[f"E{row}"] = spec_row["quantity_ton"]
        sheet[f"F{row}"] = spec_row.get("vendor", vendor)

    sheet["C27"] = today
    sheet["C27"].alignment = LEFT_ALIGNMENT
    sheet["H27"] = today
    sheet["C28"] = f" {sender}    (인)"
    sheet["H28"] = f" {receiver}    (인)"

    # 품질검사 체크리스트의 시공담당자/담당감리자 서명란 — 템플릿에는
    # 예시 이름("안진우"/"박영철")이 고정으로 박혀 있어서 실제로는 누가
    # 검사했는지와 무관하게 항상 같은 이름이 출력되고 있었다.
    sheet["C58"] = f" {checklist_sender}        (인)"
    sheet["H58"] = f" {checklist_supervisor}         (인)"

    sheet["H35"] = delivery_date
    sheet["H36"] = today_korean
    sheet["H37"] = vendor
    total_ton = round(sum(spec_row["quantity_ton"] for spec_row in specs), 3)
    sheet["H38"] = f"{_format_ton(total_ton)} Ton"
    sheet["C39"] = _build_material_spec_summary(material_type, specs)

    for row in CHECKLIST_RESULT_ROWS:
        sheet[f"G{row}"] = None

    # PHOTO_SET_ROW_START 바로 앞(사진대지 제목 행)에 명시적 페이지 나눔이
    # 없으면, 인쇄 시 어디서 페이지가 넘어갈지는 Excel이 그 PC에 설치된
    # 기본 프린터의 여백을 기준으로 자동 계산한다 — PC마다 계산이 달라져
    # "사진대지" 제목이 이전 페이지 하단에 눌려 붙는 문제가 실제로 발생했다.
    # 명시적으로 끊어 두면 어떤 PC에서 열어도 항상 이 제목부터 새 페이지가
    # 시작된다.
    sheet.row_breaks.append(Break(id=PHOTO_SET_ROW_START - 2))

    photo_sets = photo_sets or []
    non_empty_sets = [s for s in photo_sets if s.get("top") or s.get("bottom")][:MAX_PHOTO_SETS]

    for index, photo_set in enumerate(non_empty_sets):
        if index > 0:
            target_start = PHOTO_SET_ROW_START + index * PHOTO_SET_BLOCK_ROWS
            sheet.insert_rows(target_start, amount=PHOTO_SET_BLOCK_ROWS)
            _copy_photo_set_block(sheet, PHOTO_SET_ROW_START, target_start)
            sheet.row_breaks.append(Break(id=target_start - 1))
        top_anchor = PHOTO_SET_ROW_START + index * PHOTO_SET_BLOCK_ROWS
        bottom_anchor = top_anchor + 3
        sheet[f"H{top_anchor + 2}"] = delivery_date
        sheet[f"H{bottom_anchor + 2}"] = delivery_date
        report_photos.insert_photo_grid(sheet, anchor_row=top_anchor, photos=photo_set.get("top") or [])
        report_photos.insert_photo_grid(sheet, anchor_row=bottom_anchor, photos=photo_set.get("bottom") or [])

    if not non_empty_sets:
        sheet["H83"] = delivery_date
        sheet["H86"] = delivery_date

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), skipped_specs
