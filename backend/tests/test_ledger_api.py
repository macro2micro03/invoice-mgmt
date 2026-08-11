from io import BytesIO
from urllib.parse import unquote

from openpyxl import load_workbook

from fastapi.testclient import TestClient

from app.main import app

client = TestClient(app)


def _create_invoice(vendor="테스트업체", spec="SHD10", weight="1.5", delivery_date="2026-04-20", item_name="철근"):
    response = client.post(
        "/invoices",
        data={
            "material_type": "철근",
            "vendor": vendor,
            "delivery_date": delivery_date,
            "item_name": item_name,
            "spec": spec,
            "unit": "Ton",
            "weight": weight,
            "quantity": weight,
        },
    )
    return response.json()["id"]


def test_ledger_endpoint_fills_rebar_sheet_and_returns_xlsx():
    id1 = _create_invoice(spec="SHD10", weight="1.5", delivery_date="2026-04-20")
    id2 = _create_invoice(spec="SHD13", weight="2.75", delivery_date="2026-04-21")

    response = client.post(
        "/reports/material-ledger",
        data={"invoice_ids": f"{id1},{id2}", "inspector": "김검수", "supervisor": "박감리"},
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    sheet = wb["철근"]
    assert sheet["D7"].value == "SHD10"
    assert sheet["G7"].value == 1.5
    assert sheet["D8"].value == "SHD13"
    assert sheet["Q7"].value == "김검수"


def test_ledger_endpoint_accumulates_across_multiple_generations():
    id1 = _create_invoice(spec="SHD10", weight="1.0", delivery_date="2026-06-01")
    response = client.post(
        "/reports/material-ledger", data={"invoice_ids": str(id1), "inspector": "김검수", "supervisor": "박감리"}
    )
    assert response.status_code == 200

    id2 = _create_invoice(spec="SHD13", weight="2.0", delivery_date="2026-06-02")
    # id1은 다시 선택해도 이미 포함되어 있으므로 건너뛰고, id2만 새로 추가된다.
    response = client.post(
        "/reports/material-ledger", data={"invoice_ids": f"{id1},{id2}", "inspector": "이검수", "supervisor": "최감리"}
    )
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    sheet = wb["철근"]
    assert sheet["D7"].value == "SHD10"
    assert sheet["Q7"].value == "김검수"  # 처음 생성 시 값 유지, 덮어쓰지 않음
    assert sheet["D8"].value == "SHD13"
    assert sheet["Q8"].value == "이검수"

    warnings_header = response.headers.get("x-report-warnings")
    assert warnings_header is not None
    assert "1건" in unquote(warnings_header)


def test_ledger_endpoint_excludes_coupler_and_warns():
    rebar_id = _create_invoice(spec="SHD10", weight="1.0", item_name="철근")
    coupler_id = _create_invoice(spec="SHD10", weight="1.0", item_name="커플러")

    response = client.post(
        "/reports/material-ledger",
        data={"invoice_ids": f"{rebar_id},{coupler_id}"},
    )
    assert response.status_code == 200
    warnings_header = response.headers.get("x-report-warnings")
    assert warnings_header is not None
    assert "1건" in unquote(warnings_header)


def test_ledger_endpoint_400_when_nothing_to_include():
    coupler_id = _create_invoice(spec="SHD10", weight="1.0", item_name="커플러")
    response = client.post(
        "/reports/material-ledger",
        data={"invoice_ids": str(coupler_id)},
    )
    assert response.status_code == 400


def test_ledger_endpoint_400_when_invoice_ids_missing():
    response = client.post("/reports/material-ledger", data={})
    assert response.status_code == 400


def test_ledger_endpoint_is_protected_by_shared_password(monkeypatch):
    from app import config

    invoice_id = _create_invoice()
    monkeypatch.setattr(config, "APP_PASSWORD", "secret123")
    response = client.post(
        "/reports/material-ledger",
        data={"invoice_ids": str(invoice_id)},
    )
    assert response.status_code == 401


def test_get_ledger_entries_returns_sorted_list():
    id1 = _create_invoice(spec="SHD13", weight="1.0", delivery_date="2026-07-02")
    id2 = _create_invoice(spec="SHD10", weight="1.0", delivery_date="2026-07-01")
    client.post("/reports/material-ledger", data={"invoice_ids": f"{id1},{id2}"})

    response = client.get("/reports/material-ledger/entries")
    assert response.status_code == 200
    body = response.json()
    specs = [entry["spec"] for entry in body]
    assert "SHD10" in specs and "SHD13" in specs
    assert specs.index("SHD10") < specs.index("SHD13")


def test_put_ledger_entry_updates_manual_fields():
    invoice_id = _create_invoice()
    client.post("/reports/material-ledger", data={"invoice_ids": str(invoice_id)})

    response = client.put(
        f"/reports/material-ledger/entries/{invoice_id}",
        json={
            "defect_qty": 0.5,
            "defect_reason": "표면 손상",
            "release_date": "2026-05-10",
            "release_qty": 0.3,
            "remaining_qty": 0.2,
            "inspector": "김검수",
            "supervisor": "박감리",
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["defect_qty"] == 0.5
    assert body["defect_reason"] == "표면 손상"
    assert body["inspector"] == "김검수"


def test_put_ledger_entry_missing_returns_404():
    response = client.put("/reports/material-ledger/entries/999999", json={"defect_qty": 1.0})
    assert response.status_code == 404


def test_delete_ledger_entry_removes_it_and_keeps_invoice():
    invoice_id = _create_invoice()
    client.post("/reports/material-ledger", data={"invoice_ids": str(invoice_id)})

    response = client.delete(f"/reports/material-ledger/entries/{invoice_id}")
    assert response.status_code == 204

    entries = client.get("/reports/material-ledger/entries").json()
    assert all(entry["invoice_id"] != invoice_id for entry in entries)

    invoice_response = client.get(f"/invoices/{invoice_id}")
    assert invoice_response.status_code == 200


def test_delete_ledger_entry_missing_still_returns_204():
    response = client.delete("/reports/material-ledger/entries/999999")
    assert response.status_code == 204
