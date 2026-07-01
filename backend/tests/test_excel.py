from datetime import date

from openpyxl import load_workbook

from app import config, excel, models


def make_invoice(**overrides):
    defaults = dict(
        id=1,
        material_type="철근",
        vendor="대한제강",
        delivery_date=date(2026, 7, 1),
        vehicle_no="12가3456",
        invoice_no="INV-001",
        item_name="철근 D10",
        spec="D10",
        unit="TON",
        quantity=10.5,
        weight=10500,
        note="",
    )
    defaults.update(overrides)
    return models.Invoice(**defaults)


def test_append_invoice_creates_sheet_with_header_and_row():
    if config.EXCEL_PATH.exists():
        config.EXCEL_PATH.unlink()
    excel.append_invoice(make_invoice())
    workbook = load_workbook(config.EXCEL_PATH)
    assert "철근" in workbook.sheetnames
    sheet = workbook["철근"]
    assert sheet.cell(row=1, column=1).value == "id"
    assert sheet.cell(row=2, column=2).value == "대한제강"


def test_append_invoice_appends_to_existing_sheet():
    if config.EXCEL_PATH.exists():
        config.EXCEL_PATH.unlink()
    excel.append_invoice(make_invoice(id=1))
    excel.append_invoice(make_invoice(id=2, vendor="B업체"))
    workbook = load_workbook(config.EXCEL_PATH)
    sheet = workbook["철근"]
    assert sheet.max_row == 3
    assert sheet.cell(row=3, column=2).value == "B업체"


def test_append_invoice_separates_sheets_by_material_type():
    if config.EXCEL_PATH.exists():
        config.EXCEL_PATH.unlink()
    excel.append_invoice(make_invoice(id=1, material_type="철근"))
    excel.append_invoice(make_invoice(id=2, material_type="시멘트"))
    workbook = load_workbook(config.EXCEL_PATH)
    assert set(workbook.sheetnames) == {"철근", "시멘트"}


def test_append_invoice_handles_none_optional_fields():
    if config.EXCEL_PATH.exists():
        config.EXCEL_PATH.unlink()
    invoice = make_invoice(
        id=1,
        vendor=None,
        delivery_date=None,
        quantity=None,
        note=None
    )
    excel.append_invoice(invoice)
    workbook = load_workbook(config.EXCEL_PATH)
    sheet = workbook["철근"]
    assert sheet.max_row == 2  # header + 1 data row

    # Verify empty cells for None fields
    row = 2
    assert sheet.cell(row=row, column=2).value is None  # vendor
    assert sheet.cell(row=row, column=3).value is None  # delivery_date (empty string becomes None in openpyxl)
    assert sheet.cell(row=row, column=9).value is None  # quantity
    assert sheet.cell(row=row, column=11).value is None  # note

    # Verify non-None fields are correct
    assert sheet.cell(row=row, column=1).value == 1  # id
    assert sheet.cell(row=row, column=6).value == "철근 D10"  # item_name
    assert sheet.cell(row=row, column=7).value == "D10"  # spec
    assert sheet.cell(row=row, column=8).value == "TON"  # unit
    assert sheet.cell(row=row, column=10).value == 10500  # weight
