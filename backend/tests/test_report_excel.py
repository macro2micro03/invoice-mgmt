from datetime import date
from io import BytesIO
from pathlib import Path

from PIL import Image as _PILImage

from openpyxl import load_workbook

from app import report_excel

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "app" / "templates" / "material_inspection_form.xlsx"


def _photo_bytes():
    img = _PILImage.new("RGB", (100, 100), (0, 255, 0))
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _make_specs():
    return [
        {"spec": "SHD10", "quantity_ton": 3.606},
        {"spec": "SHD13", "quantity_ton": 21.11},
    ]


def _fill(**overrides):
    kwargs = dict(
        template_path=TEMPLATE_PATH,
        project_name="테스트현장 신축공사",
        work_type="건축",
        material_type="철근",
        document_number="건축(자검)-철근-1호",
        sender="김현장",
        receiver="박감리",
        checklist_sender="이시공",
        checklist_supervisor="최감리",
        specs=_make_specs(),
        vendor="동경강업(주)/동국제강",
        delivery_date="2026-03-31",
    )
    kwargs.update(overrides)
    return report_excel.fill_material_inspection_form(**kwargs)


def test_fill_material_inspection_form_sets_header_fields():
    xlsx_bytes, skipped = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["B2"].value == "테스트현장 신축공사"
    assert sheet["B4"].value == "건축(자검)-철근-1호"
    assert sheet["C28"].value == " 김현장    (인)"
    assert sheet["H28"].value == " 박감리    (인)"
    assert skipped == []


def test_fill_material_inspection_form_sets_checklist_signature_names():
    # 체크리스트의 시공담당자/담당감리자 서명란 — 예전에는 템플릿에 박힌
    # 예시 이름("안진우"/"박영철")이 실제 검사자와 무관하게 항상 그대로
    # 출력됐다.
    xlsx_bytes, _ = _fill(checklist_sender="이시공", checklist_supervisor="최감리")
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert "이시공" in sheet["C58"].value
    assert "최감리" in sheet["H58"].value
    assert "안진우" not in sheet["C58"].value
    assert "박영철" not in sheet["H58"].value


def test_fill_material_inspection_form_marks_selected_work_type_checkbox():
    xlsx_bytes, _ = _fill(work_type="토목")
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert "토목 ■" in sheet["B3"].value
    assert "건축 □" in sheet["B3"].value


def test_fill_material_inspection_form_fills_material_rows():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["A9"].value == "철근"
    assert sheet["B9"].value == "SHD10"
    assert sheet["D9"].value == "Ton"
    assert sheet["E9"].value == 3.606
    assert sheet["F9"].value == "동경강업(주)/동국제강"
    assert sheet["B10"].value == "SHD13"
    assert sheet["E10"].value == 21.11


def test_fill_material_inspection_form_uses_per_row_vendor_when_provided():
    xlsx_bytes, _ = _fill(
        specs=[
            {"spec": "SHD10", "quantity_ton": 1.0, "vendor": "동경강업(주)/동국제강"},
            {"spec": "SHD13", "quantity_ton": 0.5, "vendor": "대한제강"},
        ]
    )
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["F9"].value == "동경강업(주)/동국제강"
    assert sheet["F10"].value == "대한제강"


def test_fill_material_inspection_form_falls_back_to_top_level_vendor_when_row_has_none():
    xlsx_bytes, _ = _fill(specs=[{"spec": "SHD10", "quantity_ton": 1.0}], vendor="공통거래처")
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["F9"].value == "공통거래처"


def test_fill_material_inspection_form_computes_summary_fields():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["H35"].value == "2026-03-31"
    assert sheet["H37"].value == "동경강업(주)/동국제강"
    assert sheet["H38"].value == "24.716 Ton"
    assert sheet["C39"].value == "철근 SHD10 외 1"
    assert sheet["H83"].value == "2026-03-31"
    assert sheet["H86"].value == "2026-03-31"


def test_fill_material_inspection_form_single_spec_summary_omits_count():
    xlsx_bytes, _ = _fill(specs=[{"spec": "SHD10", "quantity_ton": 3.606}])
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["C39"].value == "철근 SHD10"


def test_fill_material_inspection_form_clears_checklist_result_column():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    for row in range(63, 80):
        assert sheet[f"G{row}"].value is None


def test_fill_material_inspection_form_leaves_inspection_result_columns_blank():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["H9"].value is None
    assert sheet["I9"].value is None
    assert sheet["J9"].value is None


def test_fill_material_inspection_form_centers_material_name_and_spec_columns():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    for row in (9, 10):
        assert sheet[f"A{row}"].alignment.horizontal == "center"
        assert sheet[f"B{row}"].alignment.horizontal == "center"
        assert sheet[f"D{row}"].alignment.horizontal == "center"


def test_fill_material_inspection_form_left_aligns_document_date():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["C27"].alignment.horizontal == "left"


def test_fill_material_inspection_form_uses_korean_date_format_for_receipt_and_inspection_dates():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    today = date.today()
    expected = f"{today:%Y}년 {today:%m}월 {today:%d}일"
    assert sheet["G4"].value == expected
    assert sheet["G5"].value == expected
    assert sheet["H36"].value == expected


def test_fill_material_inspection_form_reports_skipped_specs_beyond_capacity():
    many_specs = [{"spec": f"SPEC{i}", "quantity_ton": 1.0} for i in range(20)]
    xlsx_bytes, skipped = _fill(specs=many_specs)
    assert len(skipped) == 4
    assert skipped[0]["spec"] == "SPEC16"
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["A24"].value == "철근"
    assert sheet["B24"].value == "SPEC15"


def test_fill_material_inspection_form_inserts_top_and_bottom_photos():
    xlsx_bytes, _ = _fill(photo_sets=[{"top": [_photo_bytes(), _photo_bytes()], "bottom": [_photo_bytes()]}])
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert len(sheet._images) == 3


def test_fill_material_inspection_form_no_photos_means_no_images():
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert len(sheet._images) == 0


def test_fill_material_inspection_form_always_breaks_page_before_photo_ledger_title():
    # 80행("사 진 대 지" 제목) 앞에 명시적 페이지 나눔이 없으면, 인쇄 시
    # 어디서 페이지가 넘어갈지는 PC에 설치된 기본 프린터의 여백을 기준으로
    # Excel이 자동 계산한다 — PC마다 계산이 달라져 실제로 이 제목이 이전
    # 페이지 하단에 눌려 붙는 문제가 발생했다. 사진 세트가 없어도(사진대지가
    # 여전히 출력되므로) 항상 이 경계에 수동 페이지 나눔이 있어야 한다.
    xlsx_bytes, _ = _fill()
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["A80"].value == "사 진 대 지"
    assert any(brk.id == 79 for brk in sheet.row_breaks.brk)


def test_fill_material_inspection_form_single_set_matches_original_positions():
    xlsx_bytes, _ = _fill(photo_sets=[{"top": [_photo_bytes()], "bottom": [_photo_bytes()]}])
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert sheet["A80"].value == "사 진 대 지"
    assert sheet["H83"].value == "2026-03-31"
    assert sheet["H86"].value == "2026-03-31"
    assert len(sheet._images) == 2


def test_fill_material_inspection_form_creates_additional_rows_for_second_set():
    xlsx_bytes, _ = _fill(
        photo_sets=[
            {"top": [_photo_bytes()], "bottom": [_photo_bytes()]},
            {"top": [_photo_bytes()], "bottom": [_photo_bytes()]},
        ]
    )
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert len(sheet._images) == 4
    assert sheet["H89"].value == "2026-03-31"
    assert sheet["H92"].value == "2026-03-31"

    # Regression coverage for _copy_photo_set_block: the duplicated block
    # (rows 87-92) must carry over the same merged ranges, row heights,
    # and cell values as the source block (rows 81-86).
    coords = {m.coord for m in sheet.merged_cells.ranges}
    assert {"A87:J87", "A90:J90", "H89:J89", "C88:E88"} <= coords
    assert sheet.row_dimensions[87].height == sheet.row_dimensions[81].height
    assert sheet.row_dimensions[88].height == sheet.row_dimensions[82].height
    assert sheet["A88"].value == sheet["A82"].value == "공 종 명"

    # Finding 1: a manual row break must be inserted immediately before
    # each newly-duplicated block so printed pages don't drift out of
    # phase with the block boundaries.
    assert any(brk.id == 86 for brk in sheet.row_breaks.brk)


def test_fill_material_inspection_form_skips_empty_sets():
    xlsx_bytes, _ = _fill(
        photo_sets=[
            {"top": [], "bottom": []},
            {"top": [_photo_bytes()], "bottom": []},
        ]
    )
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert len(sheet._images) == 1
    assert sheet["H83"].value == "2026-03-31"


def test_fill_material_inspection_form_caps_at_five_sets():
    photo_sets = [{"top": [_photo_bytes()], "bottom": []} for _ in range(7)]
    xlsx_bytes, _ = _fill(photo_sets=photo_sets)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert len(sheet._images) == 5


def test_fill_material_inspection_form_empty_leading_sets_do_not_crowd_out_later_photos():
    photo_sets = [{"top": [], "bottom": []}] + [
        {"top": [_photo_bytes()], "bottom": []} for _ in range(6)
    ]
    xlsx_bytes, _ = _fill(photo_sets=photo_sets)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb.active
    assert len(sheet._images) == 5
