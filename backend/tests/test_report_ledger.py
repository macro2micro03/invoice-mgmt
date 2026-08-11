from datetime import date
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from app import report_ledger

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "app" / "templates" / "material_ledger.xlsx"


def _entry(spec, weight, delivery_date, **overrides):
    invoice = SimpleNamespace(delivery_date=delivery_date, spec=spec, weight=weight)
    defaults = dict(
        invoice=invoice,
        defect_qty=None,
        defect_reason=None,
        release_date=None,
        release_qty=None,
        remaining_qty=None,
        inspector=None,
        supervisor=None,
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def test_fill_material_ledger_writes_rows_in_order_starting_at_row_7():
    entries = [
        _entry("SHD10", 1.5, date(2026, 4, 20)),
        _entry("SHD13", 2.75, date(2026, 4, 21)),
    ]
    xlsx_bytes = report_ledger.fill_material_ledger(TEMPLATE_PATH, entries)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb["철근"]

    assert sheet["B7"].value == 1
    assert sheet["C7"].value.date() == date(2026, 4, 20)
    assert sheet["D7"].value == "SHD10"
    assert sheet["G7"].value == 1.5

    assert sheet["B8"].value == 2
    assert sheet["C8"].value.date() == date(2026, 4, 21)
    assert sheet["D8"].value == "SHD13"
    assert sheet["G8"].value == 2.75


def test_fill_material_ledger_writes_manual_fields():
    entries = [
        _entry(
            "SHD10",
            1.5,
            date(2026, 4, 20),
            defect_qty=0.2,
            defect_reason="표면 손상",
            release_date=date(2026, 5, 1),
            release_qty=1.0,
            remaining_qty=0.3,
            inspector="김검수",
            supervisor="박감리",
        )
    ]
    xlsx_bytes = report_ledger.fill_material_ledger(TEMPLATE_PATH, entries)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb["철근"]

    assert sheet["J7"].value == 0.2
    assert sheet["K7"].value == "표면 손상"
    assert sheet["N7"].value.date() == date(2026, 5, 1)
    assert sheet["O7"].value == 1.0
    assert sheet["P7"].value == 0.3
    assert sheet["Q7"].value == "김검수"
    assert sheet["R7"].value == "박감리"


def test_fill_material_ledger_leaves_manual_fields_blank_when_none():
    entries = [_entry("SHD10", 1.0, date(2026, 4, 20))]
    xlsx_bytes = report_ledger.fill_material_ledger(TEMPLATE_PATH, entries)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb["철근"]

    assert sheet["J7"].value is None
    assert sheet["K7"].value is None
    assert sheet["Q7"].value is None


def test_fill_material_ledger_preserves_existing_formulas():
    entries = [_entry("SHD10", 1.0, date(2026, 4, 20))]
    xlsx_bytes = report_ledger.fill_material_ledger(TEMPLATE_PATH, entries)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb["철근"]

    assert sheet["F7"].value == "=G7"
    assert sheet["H7"].value == '=IF(G7="","",(G7-J7))'


def test_fill_material_ledger_does_not_touch_coupler_sheet():
    entries = [_entry("SHD10", 1.0, date(2026, 4, 20))]
    xlsx_bytes = report_ledger.fill_material_ledger(TEMPLATE_PATH, entries)
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb["커플러"]

    assert sheet["B7"].value is None


def test_fill_material_ledger_empty_entries_writes_no_rows():
    xlsx_bytes = report_ledger.fill_material_ledger(TEMPLATE_PATH, [])
    wb = load_workbook(BytesIO(xlsx_bytes))
    sheet = wb["철근"]

    assert sheet["B7"].value is None
