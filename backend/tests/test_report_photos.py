from io import BytesIO

from openpyxl import Workbook
from PIL import Image

from app import report_photos


def _make_test_image_bytes(width=200, height=100, color=(255, 0, 0)):
    img = Image.new("RGB", (width, height), color)
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return buffer.getvalue()


def test_compute_grid_matches_expected_dimensions():
    assert report_photos.compute_grid(1) == (1, 1)
    assert report_photos.compute_grid(2) == (2, 1)
    assert report_photos.compute_grid(3) == (2, 2)
    assert report_photos.compute_grid(4) == (2, 2)
    assert report_photos.compute_grid(5) == (3, 2)
    assert report_photos.compute_grid(6) == (3, 2)


def test_insert_photo_grid_adds_correct_number_of_images():
    wb = Workbook()
    sheet = wb.active
    photos = [_make_test_image_bytes() for _ in range(3)]
    report_photos.insert_photo_grid(sheet, anchor_row=81, photos=photos)
    assert len(sheet._images) == 3


def test_insert_photo_grid_does_nothing_when_no_photos():
    wb = Workbook()
    sheet = wb.active
    report_photos.insert_photo_grid(sheet, anchor_row=81, photos=[])
    assert len(sheet._images) == 0


def test_insert_photo_grid_preserves_aspect_ratio_within_cell_bounds():
    wb = Workbook()
    sheet = wb.active
    photos = [_make_test_image_bytes(width=800, height=200)]
    report_photos.insert_photo_grid(sheet, anchor_row=81, photos=photos)
    image = sheet._images[0]
    assert image.width <= report_photos.BLOCK_WIDTH_PX
    assert image.height <= report_photos.BLOCK_HEIGHT_PX
    assert abs((image.width / image.height) - (800 / 200)) < 0.05
