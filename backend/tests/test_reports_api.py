from datetime import date
from urllib.parse import unquote

from openpyxl import load_workbook
from io import BytesIO

from fastapi.testclient import TestClient

from app import ocr as ocr_module
from app.main import app

client = TestClient(app)


def _table_html(headers, rows):
    thead = "<tr>" + "".join(f"<td>{h}</td>" for h in headers) + "</tr>"
    tbody = "".join("<tr>" + "".join(f"<td>{c}</td>" for c in row) + "</tr>" for row in rows)
    return f"<table><thead>{thead}</thead><tbody>{tbody}</tbody></table>"


def _cover_response(spec_weight_pairs, delivery_date="2026-03-31"):
    material_rows = [[spec, "1.000", "3", str(weight), "0", "동국제강"] for spec, weight in spec_weight_pairs]
    total = sum(weight for _, weight in spec_weight_pairs)
    table_html = _table_html(
        ["철근경", "가공중량,Ton", "할증(%)", "로스감안중량,Ton", "커플러", "비고"],
        material_rows + [["계", str(total), "", str(total), "", ""]],
    )
    info_html = f"<p>납품일: {delivery_date}<br>공장명: 가짜상사(주)</p>"
    return {
        "elements": [
            {"page": 1, "category": "heading1", "content": {"html": "<h1>철근 납품 확인서</h1>", "text": ""}},
            {"page": 1, "category": "table", "content": {"html": table_html, "text": ""}},
            {"page": 1, "category": "paragraph", "content": {"html": info_html, "text": ""}},
        ]
    }


def _cover_response_no_vendor(spec_weight_pairs):
    material_rows = [[spec, "1.000", "3", str(weight), "0", "동국제강"] for spec, weight in spec_weight_pairs]
    total = sum(weight for _, weight in spec_weight_pairs)
    table_html = _table_html(
        ["철근경", "가공중량,Ton", "할증(%)", "로스감안중량,Ton", "커플러", "비고"],
        material_rows + [["계", str(total), "", str(total), "", ""]],
    )
    return {
        "elements": [
            {"page": 1, "category": "heading1", "content": {"html": "<h1>철근 납품 확인서</h1>", "text": ""}},
            {"page": 1, "category": "table", "content": {"html": table_html, "text": ""}},
        ]
    }


def _cover_response_no_table():
    return {
        "elements": [
            {"page": 1, "category": "heading1", "content": {"html": "<h1>철근 납품 확인서</h1>", "text": ""}},
            {"page": 1, "category": "paragraph", "content": {"html": "<p>공장명: 가짜상사(주)</p>", "text": ""}},
        ]
    }


def _form_fields():
    return {
        "project_name": "테스트현장 신축공사",
        "work_type": "건축",
        "material_type": "철근",
        "sender": "김현장",
        "receiver": "박감리",
    }


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_create_report_returns_xlsx(monkeypatch):
    monkeypatch.setattr(
        ocr_module, "call_upstage_ocr", lambda image_bytes, filename="x": _cover_response([("SHD10", 0.544)])
    )

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files={"files": ("cover.jpg", b"fake-image-bytes", "image/jpeg")},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet["B2"].value == "테스트현장 신축공사"


def test_create_report_400_when_no_cover_page_found(monkeypatch):
    monkeypatch.setattr(ocr_module, "call_upstage_ocr", lambda image_bytes, filename="x": {"elements": []})

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files={"files": ("random.jpg", b"fake-image-bytes", "image/jpeg")},
    )
    assert response.status_code == 400
    assert "철근 납품 확인서" in response.json()["detail"]


def test_create_report_aggregates_multiple_uploaded_files(monkeypatch):
    responses = [
        _cover_response([("SHD10", 0.675)]),
        _cover_response([("SHD10", 2.931)]),
    ]
    call_count = {"n": 0}

    def fake_ocr(image_bytes, filename="x"):
        result = responses[call_count["n"]]
        call_count["n"] += 1
        return result

    monkeypatch.setattr(ocr_module, "call_upstage_ocr", fake_ocr)

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files=[
            ("files", ("cover1.jpg", b"fake-1", "image/jpeg")),
            ("files", ("cover2.jpg", b"fake-2", "image/jpeg")),
        ],
    )
    assert response.status_code == 200
    assert call_count["n"] == 2


def test_create_report_is_protected_by_shared_password(monkeypatch):
    from app import config

    monkeypatch.setattr(config, "APP_PASSWORD", "secret123")
    monkeypatch.setattr(
        ocr_module, "call_upstage_ocr", lambda image_bytes, filename="x": _cover_response([("SHD10", 0.544)])
    )

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files={"files": ("cover.jpg", b"fake-image-bytes", "image/jpeg")},
    )
    assert response.status_code == 401


def test_create_report_content_disposition_korean_filename(monkeypatch):
    monkeypatch.setattr(
        ocr_module, "call_upstage_ocr", lambda image_bytes, filename="x": _cover_response([("SHD10", 0.544)])
    )

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files={"files": ("cover.jpg", b"fake-image-bytes", "image/jpeg")},
    )
    assert response.status_code == 200

    content_disposition = response.headers.get("content-disposition", "")

    assert "filename*=UTF-8''" in content_disposition
    assert 'filename="report.xlsx"' in content_disposition

    encoded_part = content_disposition.split("filename*=UTF-8''")[1]
    decoded_filename = unquote(encoded_part)

    today = date.today()
    assert decoded_filename.startswith(f"자재검수요청서_{today:%y%m%d}_")
    assert decoded_filename.endswith(".xlsx")


def test_create_report_no_warning_header_when_everything_parses_cleanly(monkeypatch):
    monkeypatch.setattr(
        ocr_module, "call_upstage_ocr", lambda image_bytes, filename="x": _cover_response([("SHD10", 0.544)])
    )

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files={"files": ("cover.jpg", b"fake-image-bytes", "image/jpeg")},
    )
    assert response.status_code == 200
    assert "x-report-warnings" not in response.headers


def test_create_report_warns_when_vendor_not_recognized(monkeypatch):
    monkeypatch.setattr(
        ocr_module,
        "call_upstage_ocr",
        lambda image_bytes, filename="x": _cover_response_no_vendor([("SHD10", 0.544)]),
    )

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files={"files": ("cover.jpg", b"fake-image-bytes", "image/jpeg")},
    )
    assert response.status_code == 200

    warnings_header = response.headers.get("x-report-warnings")
    assert warnings_header is not None
    decoded = unquote(warnings_header)
    assert "거래처" in decoded


def test_create_report_warns_when_pages_skipped(monkeypatch):
    responses = [
        _cover_response([("SHD10", 0.544)]),
        _cover_response_no_table(),
    ]
    call_count = {"n": 0}

    def fake_ocr(image_bytes, filename="x"):
        result = responses[call_count["n"]]
        call_count["n"] += 1
        return result

    monkeypatch.setattr(ocr_module, "call_upstage_ocr", fake_ocr)

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files=[
            ("files", ("cover1.jpg", b"fake-1", "image/jpeg")),
            ("files", ("cover2.jpg", b"fake-2", "image/jpeg")),
        ],
    )
    assert response.status_code == 200

    warnings_header = response.headers.get("x-report-warnings")
    assert warnings_header is not None
    decoded = unquote(warnings_header)
    assert "1개 페이지" in decoded


def test_create_report_warns_when_delivery_date_not_found(monkeypatch):
    responses = [_cover_response_no_vendor([("SHD10", 0.544)])]

    def fake_ocr(image_bytes, filename="x"):
        return responses[0]

    monkeypatch.setattr(ocr_module, "call_upstage_ocr", fake_ocr)

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files={"files": ("cover.jpg", b"fake-image-bytes", "image/jpeg")},
    )
    assert response.status_code == 200

    warnings_header = response.headers.get("x-report-warnings")
    assert warnings_header is not None
    decoded = unquote(warnings_header)
    assert "반입일자" in decoded


def test_create_report_accepts_photo_uploads(monkeypatch):
    from io import BytesIO as _BytesIO

    from PIL import Image as _PILImage

    def _photo_bytes():
        img = _PILImage.new("RGB", (100, 100), (0, 255, 0))
        buf = _BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    monkeypatch.setattr(
        ocr_module, "call_upstage_ocr", lambda image_bytes, filename="x": _cover_response([("SHD10", 0.544)])
    )

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files=[
            ("files", ("cover.jpg", b"fake-image-bytes", "image/jpeg")),
            ("photo_set_1_top", ("top1.png", _photo_bytes(), "image/png")),
            ("photo_set_1_bottom", ("bottom1.png", _photo_bytes(), "image/png")),
        ],
    )
    assert response.status_code == 200
    workbook = load_workbook(_BytesIO(response.content))
    sheet = workbook.active
    assert len(sheet._images) == 2


def test_create_report_accepts_multiple_photo_sets(monkeypatch):
    from io import BytesIO as _BytesIO

    from PIL import Image as _PILImage

    def _photo_bytes():
        img = _PILImage.new("RGB", (100, 100), (0, 255, 0))
        buf = _BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()

    monkeypatch.setattr(
        ocr_module, "call_upstage_ocr", lambda image_bytes, filename="x": _cover_response([("SHD10", 0.544)])
    )

    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
        files=[
            ("files", ("cover.jpg", b"fake-image-bytes", "image/jpeg")),
            ("photo_set_1_top", ("s1top.png", _photo_bytes(), "image/png")),
            ("photo_set_2_top", ("s2top.png", _photo_bytes(), "image/png")),
            ("photo_set_2_bottom", ("s2bottom.png", _photo_bytes(), "image/png")),
        ],
    )
    assert response.status_code == 200
    workbook = load_workbook(_BytesIO(response.content))
    sheet = workbook.active
    assert len(sheet._images) == 3


def test_create_report_from_delivery_date_returns_xlsx(monkeypatch):
    from app import excel as excel_module
    from app import pdf as pdf_module

    monkeypatch.setattr(excel_module, "append_invoice", lambda invoice: None)
    monkeypatch.setattr(pdf_module, "generate_pdf", lambda invoice: "pdf/x.pdf")

    client.post(
        "/invoices",
        data={
            "material_type": "철근",
            "vendor": "동경강업(주)",
            "delivery_date": "2026-04-20",
            "spec": "SHD10",
            "weight": "1000",
            "note": "동국제강",
        },
    )
    client.post(
        "/invoices",
        data={
            "material_type": "철근",
            "vendor": "대한제강",
            "delivery_date": "2026-04-20",
            "spec": "SHD13",
            "weight": "500",
            "note": "",
        },
    )

    response = client.post(
        "/reports/material-inspection",
        data={**_form_fields(), "delivery_date": "2026-04-20"},
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet["A9"].value == "철근"
    assert sheet["F9"].value == "동경강업(주)/동국제강"
    assert sheet["F10"].value == "대한제강"
    assert sheet["H35"].value == "2026-04-20"


def test_create_report_from_delivery_date_400_when_no_records():
    response = client.post(
        "/reports/material-inspection",
        data={**_form_fields(), "delivery_date": "2099-01-01"},
    )
    assert response.status_code == 400
    assert "철근 기록이 없습니다" in response.json()["detail"]


def test_create_report_from_delivery_date_forces_material_type_to_rebar(monkeypatch):
    from app import excel as excel_module
    from app import pdf as pdf_module

    monkeypatch.setattr(excel_module, "append_invoice", lambda invoice: None)
    monkeypatch.setattr(pdf_module, "generate_pdf", lambda invoice: "pdf/x.pdf")

    client.post(
        "/invoices",
        data={
            "material_type": "철근",
            "vendor": "동경강업(주)",
            "delivery_date": "2026-05-15",
            "spec": "SHD10",
            "weight": "1000",
            "note": "동국제강",
        },
    )

    response = client.post(
        "/reports/material-inspection",
        data={**_form_fields(), "material_type": "콘크리트", "delivery_date": "2026-05-15"},
    )
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet["B4"].value.startswith("건축(자검) - 철근 - ")
    assert sheet["A9"].value == "철근"
    assert sheet["C39"].value.startswith("철근")


def test_create_report_400_when_delivery_date_malformed():
    response = client.post(
        "/reports/material-inspection",
        data={**_form_fields(), "delivery_date": "not-a-date"},
    )
    assert response.status_code == 400
    assert response.json()["detail"] == "반입일자 형식이 올바르지 않습니다 (YYYY-MM-DD)"


def test_create_report_400_when_neither_files_nor_delivery_date_given():
    response = client.post(
        "/reports/material-inspection",
        data=_form_fields(),
    )
    assert response.status_code == 400
    assert "파일을 업로드하거나 반입일자를 선택" in response.json()["detail"]


def test_create_report_from_invoice_ids_returns_xlsx(monkeypatch):
    from app import excel as excel_module
    from app import pdf as pdf_module

    monkeypatch.setattr(excel_module, "append_invoice", lambda invoice: None)
    monkeypatch.setattr(pdf_module, "generate_pdf", lambda invoice: "pdf/x.pdf")

    created = []
    for vendor, spec, weight in [("동경강업(주)", "SHD10", "1000"), ("대한제강", "SHD13", "500")]:
        response = client.post(
            "/invoices",
            data={
                "material_type": "철근",
                "vendor": vendor,
                "delivery_date": "2026-04-20",
                "spec": spec,
                "weight": weight,
                "note": "",
            },
        )
        created.append(response.json()["id"])

    response = client.post(
        "/reports/material-inspection",
        data={**_form_fields(), "invoice_ids": ",".join(str(i) for i in created)},
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet["F9"].value == "동경강업(주)"
    assert sheet["F10"].value == "대한제강"
    assert sheet["H35"].value == "2026-04-20"


def test_create_report_from_invoice_ids_400_when_none_found():
    response = client.post(
        "/reports/material-inspection",
        data={**_form_fields(), "invoice_ids": "999999,999998"},
    )
    assert response.status_code == 400
    assert "선택한 송장 기록을 찾을 수 없습니다" in response.json()["detail"]


def test_create_report_from_invoice_ids_takes_precedence_over_delivery_date(monkeypatch):
    from app import excel as excel_module
    from app import pdf as pdf_module

    monkeypatch.setattr(excel_module, "append_invoice", lambda invoice: None)
    monkeypatch.setattr(pdf_module, "generate_pdf", lambda invoice: "pdf/x.pdf")

    response = client.post(
        "/invoices",
        data={
            "material_type": "철근",
            "vendor": "동경강업(주)",
            "delivery_date": "2026-04-20",
            "spec": "SHD10",
            "weight": "1000",
            "note": "",
        },
    )
    invoice_id = response.json()["id"]

    response = client.post(
        "/reports/material-inspection",
        data={**_form_fields(), "invoice_ids": str(invoice_id), "delivery_date": "2099-01-01"},
    )
    assert response.status_code == 200
